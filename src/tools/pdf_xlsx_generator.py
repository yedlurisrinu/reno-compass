"""Document generator tool (PDF and XLSX).

This module provides functions for generating human-readable planning PDFs
(with embedded dossier.json state attachments for seamless restore) and
materials list spreadsheets, both embedded with mandatory disclaimer banners.
"""

import base64
import io
import logging
from datetime import datetime

import openpyxl
from fpdf import FPDF
from pypdf import PdfReader

from domain.dossier import Dossier

logger = logging.getLogger("reno_project")


class RenoCompassPDF(FPDF):
    """Custom FPDF2 class implementing the required supportability disclaimer."""

    def header(self):
        """Header displayed on every page of the planning document."""
        self.set_font("helvetica", "I", 8)
        self.set_text_color(180, 50, 50)  # Red warning tone
        self.cell(
            0,
            10,
            "Confidential: Educational planning artifact only. Not intended for contractor distribution.",
            align="C",
            new_x="LMARGIN",
            new_y="NEXT",
        )
        self.ln(2)

    def footer(self):
        """Footer showing page number."""
        self.set_y(-15)
        self.set_font("helvetica", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()}", align="C")


def _latin1(text: str) -> str:
    """Sanitizes text for the core (latin-1) PDF fonts to avoid encoding crashes."""
    return str(text).encode("latin-1", "replace").decode("latin-1")


def _heading(pdf: "RenoCompassPDF", title: str) -> None:
    pdf.ln(2)
    pdf.set_font("helvetica", "B", 12)
    pdf.set_text_color(0, 0, 0)
    pdf.multi_cell(0, 8, _latin1(title), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 10)


def _line(pdf: "RenoCompassPDF", text: str, indent: float = 0.0) -> None:
    pdf.set_font("helvetica", "", 10)
    pdf.set_text_color(0, 0, 0)
    if indent:
        pdf.set_x(pdf.l_margin + indent)
    pdf.multi_cell(0, 5.5, _latin1(text), new_x="LMARGIN", new_y="NEXT")


def generate_dossier_pdf(dossier: Dossier) -> str:
    """Generates the safety-forward planning PDF with an embedded dossier.json attachment.

    Section order follows the consolidation-summary spec (SI-27 / CL-76): summary +
    chosen design, then SAFETY & PERMIT prominently near the top, then budget,
    logistics/displacement, quote audit, DIY, advisory checklist, phase checklists
    (iff a design was accepted), and the budget-gap bridge at the very end (iff a gap).
    The materials xlsx ships as a SEPARATE artifact and is not referenced here.

    Args:
        dossier: The session state dossier.

    Returns:
        Base64-encoded string of the generated PDF.
    """
    logger.info(f"Generating planning document PDF for session: {dossier.envelope.dossier_id}")
    pdf = RenoCompassPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    project = dossier.project
    scope = project.scope
    design = project.design
    safety = project.safety_permit
    logistics = project.logistics_feasibility
    contractor = project.contractor_validation
    diy = project.diy_planning
    synthesis = project.synthesis

    # Title block
    pdf.set_font("helvetica", "B", 16)
    pdf.multi_cell(0, 10, "RENO COMPASS BLUEPRINT PLAN", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 10)
    _line(pdf, f"Dossier ID: {dossier.envelope.dossier_id}")
    _line(pdf, f"Generated At: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}")
    _line(pdf, f"Schema Version: {dossier.envelope.schema_version}")

    # 1. Summary + chosen design
    _heading(pdf, "1. Project Summary")
    if scope:
        _line(pdf, f"Project: {scope.project_title} ({scope.project_type})")
        _line(pdf, f"Goal: {scope.stated_goal}")
        if scope.property_context:
            pc = scope.property_context
            _line(
                pdf,
                f"Property: Zip {pc.zipcode}, {pc.dwelling_type.replace('_', ' ')}, "
                f"{pc.renovation_area:g} sqft",
            )
    if design and design.chosen_design:
        cd = design.chosen_design
        _line(
            pdf,
            f"Chosen design: {cd.option_role.replace('_', ' ')} "
            f"(est. ${cd.refined_estimate.low:,.0f}-${cd.refined_estimate.high:,.0f})",
        )
    if not scope and not (design and design.chosen_design):
        _line(pdf, "Not started/defined.")

    # 2. SAFETY & PERMIT — prominent, near the top; findings carry through unchanged
    _heading(pdf, "2. Safety & Permit (review this first)")
    if safety and safety.classifications:
        for c in safety.classifications:
            _line(pdf, f"- {c.item}: {c.tier.replace('_', ' ').upper()} (source: {c.source})")
            if c.tier == "tier_1_professional":
                _line(pdf, "Requires a licensed professional. Verify with your AHJ.", indent=4)
        if safety.permit_required:
            _line(pdf, "Permits are required. Confirm the full scope with your local AHJ.")
    else:
        _line(pdf, "No safety classifications recorded.")

    # 3. Budget (ranges + verify-locally disclaimer, SI-15)
    _heading(pdf, "3. Budget")
    if scope:
        _line(pdf, f"Target: ${scope.budget_target:,.0f}   Ceiling: ${scope.budget_ceiling:,.0f}")
        if scope.ballpark_estimate:
            be = scope.ballpark_estimate
            _line(
                pdf,
                f"Ballpark: ${be.low:,.0f}-${be.high:,.0f} "
                f"(+ contingency ${be.contingency.low:,.0f}-${be.contingency.high:,.0f})",
            )
        if scope.budget_reality_check:
            _line(pdf, f"Reality check: {scope.budget_reality_check.stated_vs_ballpark}")
    _line(pdf, "All costs are estimated ranges - verify locally with real quotes.")

    # 4. Logistics & displacement
    _heading(pdf, "4. Logistics & Displacement")
    if logistics:
        disruption = logistics.disruption or {}
        _line(pdf, f"Can live through it: {disruption.get('can_live_through_it')}")
        if logistics.chosen_displacement:
            _line(pdf, f"Displacement: {logistics.chosen_displacement}")
        _line(pdf, f"Feasibility verdict: {logistics.verdict}")
    else:
        _line(pdf, "Not assessed.")

    # 5. Contractor quote audit (only if a quote was audited)
    if contractor and (contractor.coverage_check or contractor.corner_cutting_flags):
        _heading(pdf, "5. Contractor Quote Audit")
        for cov in contractor.coverage_check:
            _line(pdf, f"- {cov.required_item}: {'present' if cov.present_in_quote else 'MISSING'}")
        for flag in contractor.corner_cutting_flags:
            _line(pdf, f"- FLAG [{flag.severity}]: {flag.flag}")

    # 6. DIY plan (only if applicable). Split by the per-item decision: items the
    # family will self-perform get the full procedure (steps, hold-points, tools);
    # items they handed back are surfaced as additions to the contractor's scope so
    # nothing is silently dropped.
    if diy and diy.procedures:
        self_perform = [p for p in diy.procedures if p.user_feasible is True]
        handed_off = [p for p in diy.procedures if p.user_feasible is False]

        if self_perform:
            _heading(pdf, "6. Your DIY Plan")
            for proc in self_perform:
                _line(pdf, f"- {proc.item} ({proc.tier.replace('_', ' ')})")
                for i, step in enumerate(proc.steps or [], start=1):
                    _line(pdf, f"{i}. {step}", indent=4)
                for hp in proc.hold_points or []:
                    _line(pdf, f"Hold point: {hp}", indent=4)
                for tool in proc.tools or []:
                    note = f" ({tool.rent_or_buy_note})" if tool.rent_or_buy_note else ""
                    _line(pdf, f"Tool: {tool.tool} - {tool.purpose}{note}", indent=4)
            for tool in diy.tools_required:
                _line(pdf, f"Tool: {tool.tool} - {tool.purpose}", indent=4)

        if handed_off:
            _heading(pdf, "6b. Add to Your Contractor's Scope")
            _line(pdf, "Eligible work you chose to hand to a professional:")
            for proc in handed_off:
                _line(pdf, f"- {proc.item} ({proc.tier.replace('_', ' ')})", indent=4)

    # 7. Advisory checklist (always carried through when present)
    if contractor and contractor.advisory_checklist:
        _heading(pdf, "7. Advisory Checklist")
        for advisory in contractor.advisory_checklist:
            _line(pdf, f"- {advisory}")

    # 8. Phase checklists (iff a design was accepted)
    if synthesis and synthesis.phase_checklists:
        checklists = synthesis.phase_checklists
        _heading(pdf, "8. Phase Checklists")
        for label, items in (
            ("Before demolition", checklists.before_demolition),
            ("After demolition", checklists.after_demolition),
            ("While in progress", checklists.while_reno_in_progress),
            ("Wrap-up", checklists.wrap_up),
        ):
            _line(pdf, f"{label}:")
            for item in items:
                _line(pdf, f"- {item}", indent=4)

    # 9. Budget-gap bridge — at the very END (iff a gap); framed as an on-ramp
    if synthesis and synthesis.budget_gap_bridge:
        bridge = synthesis.budget_gap_bridge
        _heading(pdf, "9. Budget Gap - Your On-Ramp")
        _line(
            pdf,
            f"Gap to bridge: ${bridge.gap_amount:,.0f}. This is a plan on-ramp, not a dead end.",
        )
        for opt in bridge.bridge_options:
            _line(pdf, f"- {opt}")

    # Embed the raw dossier JSON structure inside the PDF (portable restore, SI-4)
    logger.debug("Embedding raw dossier JSON metadata inside PDF attachment.")
    dossier_json = dossier.model_dump_json(indent=2)
    pdf.embed_file(bytes=dossier_json.encode("utf-8"), basename="dossier.json")

    pdf_bytes = bytes(pdf.output())
    logger.info("PDF document successfully generated and encoded to base64.")
    return base64.b64encode(pdf_bytes).decode("utf-8")


def extract_dossier_json_from_pdf(pdf_bytes: bytes) -> str:
    """Parses a planning PDF to extract the embedded dossier.json attachment.

    Enforces portable session restore (SI-4).

    Args:
        pdf_bytes: Binary bytes of the uploaded PDF file.

    Returns:
        The raw JSON string extracted from the PDF attachments.
    """
    logger.info("Extracting dossier JSON from uploaded PDF blueprint.")
    stream = io.BytesIO(pdf_bytes)
    reader = PdfReader(stream)
    attachments = reader.attachments

    if "dossier.json" not in attachments:
        logger.error(
            "Dossier extraction failed: 'dossier.json' attachment not found inside the PDF metadata."
        )
        raise ValueError(
            "No embedded 'dossier.json' attachment found in the uploaded PDF blueprint."
        )

    attachment_content = attachments["dossier.json"]
    if isinstance(attachment_content, list):
        attachment_content = attachment_content[0]

    logger.info("Dossier JSON extracted successfully from PDF attachments.")
    return attachment_content.decode("utf-8")


def generate_materials_xlsx(dossier: Dossier) -> str:
    """Generates a shoppable Excel materials list with disclaimer banner.

    Args:
        dossier: The complete session dossier.

    Returns:
        Base64-encoded string of the generated XLSX spreadsheet.
    """
    logger.info(
        f"Generating shoppable materials spreadsheet for session: {dossier.envelope.dossier_id}"
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materials Shopping List"

    # 1. Write Supportability disclaimer banner
    ws.append(
        [
            "Confidential: Educational planning artifact only. Not intended for contractor distribution."
        ]
    )
    ws.append([])  # Spacer row

    # Headers
    headers = [
        "Material Name",
        "Category",
        "Room Reference",
        "Quantity",
        "Unit",
        "Pricing Mode",
        "Waste Factor (%)",
        "Unit Cost ($)",
        "Extended Cost (Low) ($)",
        "Extended Cost (High) ($)",
        "Allergy Screened",
        "Envelope Check Status",
    ]
    ws.append(headers)

    # Style the headers briefly
    ws.row_dimensions[3].height = 24

    # 2. Write itemized materials rows
    materials_stage = dossier.project.materials
    if materials_stage and materials_stage.line_items:
        logger.debug(
            f"Writing {len(materials_stage.line_items)} material line items into spreadsheet."
        )
        for item in materials_stage.line_items:
            ext_low = item.extended_cost.get("low", 0.0)
            ext_high = item.extended_cost.get("high", 0.0)
            ws.append(
                [
                    item.material,
                    item.category,
                    item.room_ref,
                    item.quantity,
                    item.unit,
                    item.pricing_mode,
                    item.waste_factor_pct,
                    item.unit_cost,
                    ext_low,
                    ext_high,
                    "Yes" if item.allergy_screened else "No/Unscreened",
                    item.envelope_check,
                ]
            )
    else:
        logger.debug("No materials line items found to write in spreadsheet.")

    # Write total rollup summary at bottom
    if materials_stage and materials_stage.final_total:
        ws.append([])
        ws.append(["TOTAL ESTIMATED MATERIALS OUTLAY:"])
        ws.append(
            [
                f"Low range: ${materials_stage.final_total.low:,.2f}",
                f"High range: ${materials_stage.final_total.high:,.2f}",
                f"Allowance portion: ${materials_stage.final_total.allowance_portion:,.2f}",
            ]
        )

    # Save spreadsheet to bytes
    stream = io.BytesIO()
    wb.save(stream)
    excel_bytes = stream.getvalue()
    logger.info("Materials spreadsheet generated successfully and encoded to base64.")
    return base64.b64encode(excel_bytes).decode("utf-8")
