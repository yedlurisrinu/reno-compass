"""Coverage for the fully-populated PDF/XLSX rendering branches.

The base ``test_pdf_xlsx_generator`` module renders an empty dossier; this one
drives every optional section (logistics, contractor audit, DIY self-perform +
handed-off, phase checklists, budget-gap bridge, itemized materials + rollup)
plus the "no embedded attachment" extraction error.
"""

import base64
from datetime import datetime

from fpdf import FPDF

from domain.dossier import (
    BudgetGapBridge,
    ContractorValidationStage,
    CornerCuttingFlag,
    CoverageCheckItem,
    DiyPlanningStage,
    DiyProcedure,
    Dossier,
    DossierEnvelope,
    LogisticsFeasibilityStage,
    PhaseChecklists,
    ProjectBody,
    SectionStatus,
    SynthesisStage,
    ToolRequired,
)
from tests.unit.test_orchestrator_gates import make_design, make_materials, make_safety, make_scope
from tools.pdf_xlsx_generator import (
    extract_dossier_json_from_pdf,
    generate_dossier_pdf,
    generate_materials_xlsx,
)


def _rich_dossier() -> Dossier:
    logistics = LogisticsFeasibilityStage(
        status=SectionStatus(state="in_progress"),
        disruption={"can_live_through_it": False, "offline_utilities": ["water"]},
        chosen_displacement="Stay with family for 2 weeks",
        verdict="proceed",
    )
    contractor = ContractorValidationStage(
        status=SectionStatus(state="in_progress"),
        quote_provided=True,
        coverage_check=[
            CoverageCheckItem(required_item="Waterproofing", present_in_quote=True, note="ok"),
            CoverageCheckItem(required_item="Demolition", present_in_quote=False, note="missing"),
        ],
        corner_cutting_flags=[CornerCuttingFlag(flag="No permit line item", severity="high")],
        advisory_checklist=["Get 3-5 bids", "Verify license"],
    )
    diy = DiyPlanningStage(
        status=SectionStatus(state="in_progress"),
        procedures=[
            DiyProcedure(
                item="Install vanity",
                tier="tier_3_proceed",
                user_feasible=True,
                steps=["Remove old vanity", "Set new vanity"],
                hold_points=["Confirm plumbing rough-in"],
                tools=[ToolRequired(tool="Level", purpose="Alignment", rent_or_buy_note="buy")],
            ),
            DiyProcedure(
                item="Move plumbing",
                tier="tier_2_permitted",
                user_feasible=False,
            ),
        ],
        tools_required=[ToolRequired(tool="Drill", purpose="Mounting")],
    )
    synthesis = SynthesisStage(
        status=SectionStatus(state="in_progress"),
        phase_checklists=PhaseChecklists(
            before_demolition=["Shut off water"],
            after_demolition=["Inspect subfloor"],
            while_reno_in_progress=["Daily cleanup"],
            wrap_up=["Final walkthrough"],
        ),
        budget_gap_bridge=BudgetGapBridge(
            gap_amount=5000.0, bridge_options=["Phase the work", "Choose economy tile"]
        ),
    )
    return Dossier(
        envelope=DossierEnvelope(
            dossier_id="reno_s_rich",
            schema_version="1.0.0",
            created_at=datetime.utcnow(),
            last_updated_at=datetime.utcnow(),
            origin="fresh",
            current_stage="synthesis",
        ),
        project=ProjectBody(
            scope=make_scope(),
            design=make_design(),
            safety_permit=make_safety(),
            logistics_feasibility=logistics,
            contractor_validation=contractor,
            diy_planning=diy,
            materials=make_materials(),
            synthesis=synthesis,
        ),
    )


def test_pdf_renders_all_populated_sections():
    pdf_b64 = generate_dossier_pdf(_rich_dossier())
    pdf_bytes = base64.b64decode(pdf_b64)
    # Round-trips through the embedded-attachment extractor.
    extracted = extract_dossier_json_from_pdf(pdf_bytes)
    assert "reno_s_rich" in extracted


def test_xlsx_renders_line_items_and_rollup():
    xlsx_b64 = generate_materials_xlsx(_rich_dossier())
    assert isinstance(xlsx_b64, str)

    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(xlsx_b64)))
    ws = wb.active
    text = "\n".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
    )
    assert "Ceramic Tile" in text
    assert "TOTAL ESTIMATED MATERIALS OUTLAY:" in text


def test_extract_raises_without_embedded_attachment():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("helvetica", size=12)
    pdf.cell(0, 10, "no attachment here")
    plain_pdf = bytes(pdf.output())

    import pytest

    with pytest.raises(ValueError, match="dossier.json"):
        extract_dossier_json_from_pdf(plain_pdf)
